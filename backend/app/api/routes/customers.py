from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import io

from app.db.session import get_db
from app.models.customer import Customer
from app.schemas.customer import CustomerCreate, CustomerUpdate, CustomerOut
from app.core.security import get_current_user

router = APIRouter(dependencies=[Depends(get_current_user)])

@router.get("/", response_model=List[CustomerOut])
def list_customers(skip: int = 0, limit: int = 5000, db: Session = Depends(get_db)):
    return db.query(Customer).offset(skip).limit(limit).all()

@router.post("/", response_model=CustomerOut)
def create_customer(customer: CustomerCreate, db: Session = Depends(get_db)):
    db_customer = Customer(**customer.model_dump())
    db.add(db_customer)
    db.commit()
    db.refresh(db_customer)
    return db_customer

@router.post("/import/excel")
@router.post("/import/excel/")
async def import_customers_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    wb = openpyxl.load_workbook(io.BytesIO(await file.read()))
    ws: Optional[Worksheet] = wb.active

    if not ws:
        return {"message": "Empty workbook", "count": 0}

    header_rows = list(ws.iter_rows(min_row=1, max_row=1))
    if not header_rows:
        return {"message": "Empty workbook", "count": 0}

    header_row = header_rows[0]
    headers = [str(cell.value).strip().lower().replace(" ", "_") if cell.value is not None else "" for cell in header_row]
    col_map = {name: idx for idx, name in enumerate(headers) if name}

    def get_cell(row, *possible_names):
        for pname in possible_names:
            if pname in col_map:
                val = row[col_map[pname]].value
                if val is not None:
                    return val
        return None

    imported = 0
    for row in ws.iter_rows(min_row=2):
        company_name = get_cell(row, "company_name", "company", "firm_name", "firm", "client_name")
        contact_person = get_cell(row, "contact_person", "contact", "name", "person", "proprietor")
        
        # Check required fields
        if not company_name:
            continue
        
        # Fallback values if contact person or phone/address are missing
        final_contact = str(contact_person).strip() if contact_person else "Unknown Contact"
        
        phone = get_cell(row, "phone", "mobile", "phone_number", "contact_no", "contact_number")
        final_phone = str(phone).strip() if phone else "N/A"
        
        address = get_cell(row, "address", "location", "office_address", "billing_address")
        final_address = str(address).strip() if address else "N/A"

        email = get_cell(row, "email", "email_address")
        shipping_address = get_cell(row, "shipping_address", "delivery_address")
        state_name = get_cell(row, "state_name", "state")
        state_code = get_cell(row, "state_code", "code")
        gstin = get_cell(row, "gstin", "gst_no", "gst", "gst_number")

        customer = Customer(
            company_name=str(company_name).strip(),
            contact_person=final_contact,
            email=str(email).strip() if email else None,
            phone=final_phone,
            address=final_address,
            shipping_address=str(shipping_address).strip() if shipping_address else None,
            state_name=str(state_name).strip() if state_name else None,
            state_code=str(state_code).strip() if state_code else None,
            gstin=str(gstin).strip() if gstin else None,
        )
        db.add(customer)
        imported += 1

    db.commit()
    return {"message": f"Successfully imported {imported} customer(s).", "count": imported}

from pydantic import BaseModel

class CustomerImportRow(BaseModel):
    company_name: str
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    shipping_address: Optional[str] = None
    state_name: Optional[str] = None
    state_code: Optional[str] = None
    gstin: Optional[str] = None

class CustomerImportSaveRequest(BaseModel):
    customers: List[CustomerImportRow]

@router.post("/import/preview")
async def import_customers_preview(file: UploadFile = File(...)):
    wb = openpyxl.load_workbook(io.BytesIO(await file.read()))
    ws: Optional[Worksheet] = wb.active

    if not ws:
        return {"rows": [], "warnings": ["Empty workbook"]}

    header_rows = list(ws.iter_rows(min_row=1, max_row=1))
    if not header_rows:
        return {"rows": [], "warnings": ["Empty workbook"]}

    header_row = header_rows[0]
    headers = [str(cell.value).strip().lower().replace(" ", "_") if cell.value is not None else "" for cell in header_row]
    col_map = {name: idx for idx, name in enumerate(headers) if name}

    def get_cell(row, *possible_names):
        for pname in possible_names:
            if pname in col_map:
                val = row[col_map[pname]].value
                if val is not None:
                    return val
        return None

    rows = []
    warnings = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        company_name = get_cell(row, "company_name", "company", "firm_name", "firm", "client_name")
        contact_person = get_cell(row, "contact_person", "contact", "name", "person", "proprietor")
        
        # Check required fields
        if not company_name:
            warnings.append(f"Row {idx}: Missing required field 'company_name'")
            continue
            
        final_contact = str(contact_person).strip() if contact_person else "Unknown Contact"
        
        phone = get_cell(row, "phone", "mobile", "phone_number", "contact_no", "contact_number")
        final_phone = str(phone).strip() if phone else "N/A"
        
        address = get_cell(row, "address", "location", "office_address", "billing_address")
        final_address = str(address).strip() if address else "N/A"

        email = get_cell(row, "email", "email_address")
        shipping_address = get_cell(row, "shipping_address", "delivery_address")
        state_name = get_cell(row, "state_name", "state")
        state_code = get_cell(row, "state_code", "code")
        gstin = get_cell(row, "gstin", "gst_no", "gst", "gst_number")

        rows.append({
            "row_num": idx,
            "company_name": str(company_name).strip(),
            "contact_person": final_contact,
            "email": str(email).strip() if email else None,
            "phone": final_phone,
            "address": final_address,
            "shipping_address": str(shipping_address).strip() if shipping_address else None,
            "state_name": str(state_name).strip() if state_name else None,
            "state_code": str(state_code).strip() if state_code else None,
            "gstin": str(gstin).strip() if gstin else None,
        })

    return {"rows": rows, "warnings": warnings}

@router.post("/import/save")
def save_imported_customers(request: CustomerImportSaveRequest, db: Session = Depends(get_db)):
    imported = 0
    for row in request.customers:
        if not row.company_name or not row.company_name.strip():
            continue
        customer = Customer(
            company_name=row.company_name.strip(),
            contact_person=(row.contact_person or "Unknown Contact").strip(),
            email=row.email.strip() if row.email else None,
            phone=(row.phone or "N/A").strip(),
            address=(row.address or "N/A").strip(),
            shipping_address=row.shipping_address.strip() if row.shipping_address else None,
            state_name=row.state_name.strip() if row.state_name else None,
            state_code=row.state_code.strip() if row.state_code else None,
            gstin=row.gstin.strip() if row.gstin else None,
        )
        db.add(customer)
        imported += 1
    
    db.commit()
    return {"message": f"Successfully imported {imported} customer(s).", "count": imported}

@router.get("/{customer_id}", response_model=CustomerOut)
def get_customer(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer

@router.put("/{customer_id}", response_model=CustomerOut)
def update_customer(customer_id: int, customer_update: CustomerUpdate, db: Session = Depends(get_db)):
    db_customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not db_customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    for key, value in customer_update.model_dump().items():
        setattr(db_customer, key, value)
    db.commit()
    db.refresh(db_customer)
    return db_customer

@router.delete("/{customer_id}")
def delete_customer(customer_id: int, db: Session = Depends(get_db)):
    db_customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not db_customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    db.delete(db_customer)
    db.commit()
    return {"detail": "Customer deleted successfully"}
