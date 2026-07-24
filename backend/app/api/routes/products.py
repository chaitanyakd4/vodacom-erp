import io
from typing import List, Optional
import logging

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from fastapi import APIRouter, UploadFile, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.db.session import get_db
from app.models.product import Product
from app.schemas.product import ProductCreate, ProductUpdate, ProductOut
from app.core.security import get_current_user

router = APIRouter(prefix="/products", tags=["products"], dependencies=[Depends(get_current_user)])


@router.get("/", response_model=List[ProductOut])
def list_products(skip: int = 0, limit: int = 10000, db: Session = Depends(get_db)):
    return db.query(Product).offset(skip).limit(limit).all()


def auto_categorize(name: str, description: Optional[str] = None) -> str:
    name_lower = name.lower()
    desc_lower = description.lower() if description else ""
    combined = f"{name_lower} {desc_lower}"

    if any(k in combined for k in ['camera', 'cctv', 'nvr', 'dvr', 'xvr', 'dome', 'bullet', 'video recorder', 'hard disc', 'seagate', 'skyhawk', 'wd hard']):
        return 'CCTV & Recording'
    elif any(k in combined for k in ['switch', 'router', 'patch pannel', 'access point', 'ap-', 'ap ', '-ap', 'wifi', 'networking', 'print server', 'krone', 'media connector', 'fiber patch cord']):
        return 'Networking Equipment'
    elif any(k in combined for k in ['cable', 'cord', 'connector', 'wire', 'hdmi', 'vga', 'rj45', 'face plate', 'gane box', 'gang box', 'pchi cord', 'booster cable', 'cat 6']):
        return 'Cables & Connectors'
    elif any(k in combined for k in ['power supply', 'adaptor', 'adapter', 'ups', ' dc', 'mcb', 'power mex', 'component adaptor']):
        return 'Power Supplies, Adapters & UPS'
    elif any(k in combined for k in ['hooter', 'fire alarm', 'smoke detector', 'door', 'attendance', 'biomatric', 'biometric', 'headphone', 'speaker', 'mic', 'microphone', 'em lock', 'push button', 'essl']):
        return 'Building Systems'
    elif any(k in combined for k in ['phone', 'handset', 'dect', 'ale', 'beetel', 'panasonic', 'alcatel', 'analog', 'binatone', 'lexstar', 'procel', 'gt210', '4008', '4018', '4019', '4028', '4029', '4039', '4068', '8001', '8008', '8012', '8018', '8029', '8039', '8068', '8088']):
        return 'Telephone Handsets'
    elif any(k in combined for k in ['card', 'module', 'sli', 'uai', 'amix', 'apa', 'daughter board', 'pra', 'blank slots']):
        return 'EPABX Cards & Modules'
    elif any(k in combined for k in ['cabinet', 'base station', 'mounting kit', 'rack mount', 'rack mounting', 'indoor bases', 'oxo connect', 'suite evolution']):
        return 'EPABX Cabinets & Switches'
    elif any(k in combined for k in ['gateway', 'fct', 'repeater', 'sip', 'voip', 'cellular terminal', 'phone recording']):
        return 'VoIP / SIP & Gateway Equipment'
    else:
        return 'CCTV & Recording'


@router.post("/", response_model=ProductOut)
def create_product(
    product: ProductCreate,
    allow_update_existing: bool = Query(True, description="If product exists, merge stock & update price instead of duplicating"),
    db: Session = Depends(get_db)
):
    clean_name = product.name.strip()
    
    # ── Duplicate Prevention Check ──────────────────────────────────────────
    existing = db.query(Product).filter(
        func.lower(func.trim(Product.name)) == clean_name.lower()
    ).first()

    if not existing and product.sku:
        existing = db.query(Product).filter(
            func.lower(func.trim(Product.sku)) == product.sku.strip().lower()
        ).first()

    if existing:
        if allow_update_existing:
            # Smart Upsert: Update existing product instead of creating duplicate
            existing.stock_quantity += (product.stock_quantity or 0)
            if product.price > 0:
                existing.price = product.price
            if product.hsn_code:
                existing.hsn_code = product.hsn_code
            if product.description:
                existing.description = product.description
            if product.category and product.category != "Uncategorized":
                existing.category = product.category
            
            db.commit()
            db.refresh(existing)
            logging.info(f"[DEDUP] Updated existing product ID {existing.id} ({existing.name})")
            return existing
        else:
            raise HTTPException(
                status_code=400,
                detail=f"A product named '{clean_name}' already exists in inventory."
            )

    if product.category == "Uncategorized" or not product.category:
        product.category = auto_categorize(clean_name, product.description)

    db_product = Product(**product.model_dump())
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    return db_product


@router.post("/deduplicate")
def deduplicate_inventory(db: Session = Depends(get_db)):
    """
    Scans entire inventory for duplicate product names (case-insensitive).
    Merges duplicate stock quantities into a single primary record and deletes duplicate entries.
    """
    all_products = db.query(Product).order_by(Product.id.asc()).all()
    seen = {}
    merged_count = 0
    deleted_ids = []

    for prod in all_products:
        key = prod.name.strip().lower()
        if key in seen:
            primary = seen[key]
            # Combine stock quantity
            primary.stock_quantity += prod.stock_quantity
            # Prefer non-empty HSN / description / price
            if not primary.hsn_code and prod.hsn_code:
                primary.hsn_code = prod.hsn_code
            if primary.price == 0 and prod.price > 0:
                primary.price = prod.price
            
            db.delete(prod)
            deleted_ids.append(prod.id)
            merged_count += 1
        else:
            seen[key] = prod

    if merged_count > 0:
        db.commit()

    return {
        "status": "success",
        "merged_count": merged_count,
        "deleted_product_ids": deleted_ids,
        "total_unique_products": len(seen)
    }


@router.get("/{product_id}", response_model=ProductOut)
def get_product(product_id: int, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@router.put("/{product_id}", response_model=ProductOut)
def update_product(product_id: int, product_update: ProductUpdate, db: Session = Depends(get_db)):
    db_product = db.query(Product).filter(Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    for key, value in product_update.model_dump(exclude_unset=True).items():
        setattr(db_product, key, value)
    db.commit()
    db.refresh(db_product)
    return db_product


@router.delete("/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db)):
    db_product = db.query(Product).filter(Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    db.delete(db_product)
    db.commit()
    return {"detail": "Product deleted successfully"}


@router.post("/import/excel")
@router.post("/import/excel/")
async def import_from_excel(file: UploadFile, db: Session = Depends(get_db)):
    wb = openpyxl.load_workbook(io.BytesIO(await file.read()))
    ws: Optional[Worksheet] = wb.active

    if not ws:
        return {"message": "Empty workbook", "created": 0, "updated": 0}

    header_rows = list(ws.iter_rows(min_row=1, max_row=1))
    if not header_rows:
        return {"message": "Empty workbook", "created": 0, "updated": 0}

    header_row = header_rows[0]
    headers = [str(cell.value).strip().lower().replace(" ", "_") if cell.value is not None else "" for cell in header_row]
    col_map = {name: idx for idx, name in enumerate(headers) if name}

    def get_cell(row, *possible_names):
        for pname in possible_names:
            if pname in col_map:
                val = row[col_map[pname]].value
                if val is not None:
                    return val
        return None

    created_count = 0
    updated_count = 0

    for row in ws.iter_rows(min_row=2):
        raw_name = get_cell(row, "name", "product_name", "item_name", "product", "item", "service")
        if raw_name is None:
            continue

        clean_name = str(raw_name).strip()
        if not clean_name:
            continue

        description = get_cell(row, "description", "desc", "details")
        hsn_code = get_cell(row, "hsn_code", "hsn", "hsn_sac", "sac_code")
        unit = get_cell(row, "unit", "uom", "unit_of_measure")

        raw_price = get_cell(row, "price", "unit_price", "rate", "mrp", "cost")
        raw_tax = get_cell(row, "tax_rate", "gst", "gst_rate", "tax", "tax_%", "gst_%")
        raw_stock = get_cell(row, "stock_quantity", "stock", "qty", "quantity", "opening_stock")
        category = get_cell(row, "category", "type", "group")

        try:
            price = float(raw_price) if raw_price is not None else 0.0
        except (ValueError, TypeError):
            price = 0.0

        try:
            tax_rate = float(raw_tax) if raw_tax is not None else 18.0
        except (ValueError, TypeError):
            tax_rate = 18.0

        try:
            stock_quantity = int(float(raw_stock)) if raw_stock is not None else 0
        except (ValueError, TypeError):
            stock_quantity = 0

        final_category = str(category).strip() if category else auto_categorize(clean_name, str(description).strip() if description else None)

        # ── Smart Deduplication check ──────────────────────────────────────
        existing = db.query(Product).filter(
            func.lower(func.trim(Product.name)) == clean_name.lower()
        ).first()

        if existing:
            # Merge stock & update product details instead of creating a duplicate entry
            existing.stock_quantity += stock_quantity
            if price > 0:
                existing.price = price
            if tax_rate:
                existing.tax_rate = tax_rate
            if hsn_code:
                existing.hsn_code = str(hsn_code).strip()
            if description:
                existing.description = str(description).strip()
            if category and final_category != "Uncategorized":
                existing.category = final_category
            updated_count += 1
        else:
            product = Product(
                name=clean_name,
                description=str(description).strip() if description else None,
                category=final_category,
                hsn_code=str(hsn_code).strip() if hsn_code else None,
                price=price,
                tax_rate=tax_rate,
                stock_quantity=stock_quantity,
                unit=str(unit).strip() if unit else "pcs",
            )
            db.add(product)
            created_count += 1

    db.commit()
    return {
        "message": f"Import completed: {created_count} new item(s) created, {updated_count} existing item(s) updated with new stock.",
        "created": created_count,
        "updated": updated_count
    }